import pandas as pd
import os
import tempfile
from datetime import datetime
from typing import Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class NoDataInRange(Exception):
    """Raised when no rows remain after the date filter."""
    pass

def generate_clawback_report(
    input1_path: str,
    input2_path: str,
    start_date_str: str,
    end_date_str: str,
    report_title_input: Optional[str] = None
) -> Tuple[str, str]:

    #=============================================================================
    # 1) Read CSV and basic cleanup
    #=============================================================================
    df_raw = pd.read_excel(input1_path, dtype=str, skiprows=1, engine="openpyxl")
    df_raw.columns = df_raw.columns.str.strip()
    df_raw = df_raw[df_raw['Remarks'].isin(['Adj', 'Claw', 'New'])]

    # Construct Adviser Name
    df_raw['Adviser Name'] = df_raw['Adviser Forename'] + ' ' + df_raw['Adviser Surname']
    df_raw['Adviser Name'] = df_raw['Adviser Name'].str.replace(' 1', '', regex=True)

    # APE cleanup
    # df_raw['APE'] = pd.to_numeric(df_raw['APE'], errors='coerce').dropna()
    df_raw['APE'] = df_raw['APE'].str.replace(',', '', regex=True)
    # df_raw['APE'] = df_raw['APE'].str.replace(' ', '', regex=True)
    df_raw['APE'] = df_raw['APE'].str.replace(' -   ', '0')
    df_raw['APE'] = df_raw['APE'].str.replace('#DIV/0!', '0', regex=True)
    df_raw['APE'] = pd.to_numeric(df_raw['APE']).dropna()
    df_raw = df_raw[df_raw['APE'].abs() > 1]

    df_raw['Paid Date'] = pd.to_datetime(df_raw['Paid Date'], errors='coerce')
    df_raw['Start Date'] = pd.to_datetime(df_raw['Start Date'], errors='coerce')

    start_date = pd.to_datetime(start_date_str, errors="raise")
    end_date   = pd.to_datetime(end_date_str, errors="raise")
    df_raw = df_raw[(df_raw['Start Date'] >= start_date) & (df_raw['Start Date'] <= end_date)]

    if df_raw.empty:
        raise NoDataInRange(f"No data found between {start_date.date()} and {end_date.date()}")

    #=============================================================================
    # 2) NTU policies (zero-payment)
    #=============================================================================
    policy_totals = (
        df_raw
        .groupby('Policy Number')['APE']
        .sum()
        .round(2)
        .reset_index()
    )

    # Treat policies with exactly zero payment as NTU
    zero_policy_totals = policy_totals[policy_totals['APE'] == 0]
    ntu_policies = zero_policy_totals['Policy Number']

    filtered_df = df_raw[df_raw['Policy Number'].isin(ntu_policies)]
    main_df = (
        filtered_df
        .groupby('Adviser Name')['Policy Number']
        .nunique()
        .reset_index(name='NTU Count')
    )

    # Sum all positive APE values instead of taking the max per policy
    positive_ntu = filtered_df[filtered_df['APE'] > 0]  # <- ensure positive only
    ntu_payment = (
        positive_ntu
        .groupby('Adviser Name')['APE']
        .sum()  # Sum all positive APE values for each adviser
        .round(2)
        .reset_index(name='NTU APE')
    )

    main_df = main_df.merge(ntu_payment, on='Adviser Name', how='left')

    #=============================================================================
    # 3) Clawbacks
    #=============================================================================
    df_raw_no_ntu = df_raw[~df_raw['Policy Number'].isin(ntu_policies)]

    df_raw_no_ntu['Clawback'] = df_raw_no_ntu['Remarks'].str.contains('Claw|Adj', case=False, na=False).astype(int)

    clawback_count = (
        df_raw_no_ntu[df_raw_no_ntu['Clawback'] == 1]
        .groupby('Adviser Name')['Policy Number']
        .nunique()
        .reset_index(name='Clawback Count')
    )

    # STEP 4 & 5 Combined: Get total clawback APE per adviser from all negative-APE clawback entries
    clawback_payment = (
        df_raw_no_ntu[(df_raw_no_ntu['Clawback'] == 1) & (df_raw_no_ntu['APE'] < 0)]
        .groupby('Adviser Name')['APE']
        .sum()
        .round(2)
        .reset_index(name='Clawback APE')
    )

    # Make APE positive (represents amount clawed back)

    clawback_payment['Clawback APE'] *= -1

    df_raw_clawback = clawback_count.merge(clawback_payment, on='Adviser Name', how='left')
    main_df = main_df.merge(df_raw_clawback, on='Adviser Name', how='outer')
    

    #=============================================================================
    # 4) Raw issued business + counts (accounting for all positive payments)
    #=============================================================================
    # Filter out negative or zero APE values and consider all positive payments
    df_raw_positive = df_raw[df_raw['Remarks'] == 'New']  # <- include all positive APE values

    # Sum all positive APE values for each adviser and policy
    raw_total_payment = (
        df_raw_positive
        .groupby('Adviser Name')['APE']
        .sum()  # Sum all positive APE values for each adviser
        .round(2)
        .reset_index(name='APE')  # Renaming to 'Issued Business' for clarity
    )

    # Count the unique policies issued by each adviser
    raw_policy_count = (
        df_raw_positive
        .groupby('Adviser Name')['Policy Number']
        .nunique()
        .reset_index(name='Total Count')
    )

    # Merge the total payment and policy count into one DataFrame
    df_raw_agg = raw_total_payment.merge(raw_policy_count, on='Adviser Name', how='left')


    # Merge this with the main dataframe
    main_df = main_df.merge(df_raw_agg, on='Adviser Name', how='outer').fillna(0)


    print(main_df[main_df['Adviser Name'] == 'Adrian Tabus'])


    #=============================================================================
    # 5) Calculate percentages
    #=============================================================================
    main_df['NTU %'] = (
        main_df['NTU Count'].fillna(0)
        / main_df['Total Count'].replace(0, float('nan'))
    ).round(4)

    main_df['Clawback %'] = (
        main_df['Clawback Count'].fillna(0)
        / main_df['Total Count'].replace(0, float('nan'))
    ).round(4)

    main_df['Total %'] = (main_df['NTU %'] + main_df['Clawback %']).round(4)
    main_df.rename(columns={'Adviser Name': 'Adviser'}, inplace=True)

    #=============================================================================
    # 6) Merge Team Group
    #=============================================================================
    df_leave_manager = pd.read_excel(input2_path, dtype=str, engine="openpyxl")
    df_leave_manager.rename(columns={'NAME ': 'Adviser'}, inplace=True)
    df_leave_manager['Team Group'] = df_leave_manager['Team Group'].str.rstrip()

    # Merge Adviser, Team Group, and Email into main_df
    main_df = main_df.merge(
        df_leave_manager[['Adviser', 'Team Group']],
        on='Adviser',
        how='left'
    )

    # Keep only advisers that exist in the mapping file
    valid_advisers = set(df_leave_manager['Adviser'].unique())
    main_df = main_df[main_df['Adviser'].isin(valid_advisers)]




    # Reorder columns: Adviser, Team Group, Email, then everything else
    cols = ['Adviser', 'Team Group'] + [
        col for col in main_df.columns
        if col not in ['Adviser', 'Team Group']
    ]
    main_df = main_df[cols]
    

   #=============================================================================
    # 7) Final metrics and ordering
    #=============================================================================
    main_df["IN Books $"] = main_df["APE"] - main_df["Clawback APE"]
    main_df["IN Books Count"] = (
        main_df["Total Count"] - (main_df["Clawback Count"] + main_df["NTU Count"])
    )

    desired_columns = [
        "Adviser",
        "Team Group",
        "APE",
        "Clawback APE",
        "NTU APE",
        "IN Books $",
        "Total Count",
        "Clawback Count",
        "NTU Count",
        "IN Books Count",
        "Clawback %",
        "NTU %",
        "Total %"
    ]
    main_df = main_df[desired_columns]
    main_df.rename(columns={
        'APE': 'Issued Business $',
        'Clawback APE': 'Clawback $',
        'NTU APE': 'NTU $',
        'Total Count': 'Issued Policies Count'
    }, inplace=True)

    #=============================================================================
    # 8) Summary and export
    #=============================================================================
    main_df.fillna(0, inplace=True)
    main_df = main_df[main_df['Team Group'] != 'Left Adviser']

    percent_cols = ["Clawback %", "NTU %", "Total %"]
    main_df[percent_cols] *= 100

    sum_row = main_df.drop(columns=percent_cols).sum(numeric_only=True)
    avg_percent_row = main_df[percent_cols].mean()
    summary_row = pd.concat([sum_row, avg_percent_row])
    summary_df = pd.DataFrame(summary_row, columns=["Value"]).T

    print(summary_df)
    print(main_df['Issued Business $'].sum())

    # -----------------------------------------------------------
    # 9) Create the Provider Pivot (unique policy counts by adviser/provider)
    # -----------------------------------------------------------
    # Existing transformation
    # Initial transformation
    df_pivot = (
        df_raw[['Adviser Name', 'Provider', 'Policy Number']]
        .drop_duplicates()
        .merge(df_leave_manager[['Adviser', 'Team Group']],
            left_on='Adviser Name', right_on='Adviser', how='left')
    )

    df_pivot = (
        df_pivot
        .groupby(['Adviser', 'Team Group', 'Provider'])['Policy Number']
        .nunique()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Add total policies column
    df_pivot['Total Policies Sold'] = df_pivot.iloc[:, 2:].sum(axis=1)

    # Reorder: move Total Policies Sold to be the first column after identifiers
    cols = df_pivot.columns.tolist()
    # ['Adviser', 'Team Group', ..., 'Total Policies Sold']
    provider_cols = cols[2:-1]
    new_order = ['Adviser', 'Team Group', 'Total Policies Sold'] + provider_cols
    df_pivot = df_pivot[new_order]

    # Add % columns
    for col in provider_cols:
        df_pivot[f'{col} %'] = (df_pivot[col] / df_pivot['Total Policies Sold']) * 100

    # Optional: round for presentation
    df_pivot[[f'{col} %' for col in provider_cols]] = df_pivot[[f'{col} %' for col in provider_cols]].round(2)

    df_pivot = df_pivot[df_pivot['Team Group'] != 'Left Adviser']

    # -----------------------------------------------------------
    # Create the Manager Report
    # -----------------------------------------------------------

    # 1) Sort by Team Group (so each group is contiguous)
    main_df_sorted = main_df.sort_values(['Team Group','Adviser'], na_position='last')

    # 2) Group by Team Group
    grouped = main_df_sorted.groupby('Team Group', dropna=False)

    # Prepare a list of rows (dicts) for final assembly
    manager_rows = []

    # We'll keep columns in the exact same order as main_df,
    # EXCEPT we drop 'Team Group' so it won't show in the final output.
    ordered_columns = [col for col in main_df.columns if col != 'Team Group']

    first_group = True  # to avoid blank lines before the first group

    for team_group, grp_df in grouped:
        # (A) Add 3 blank lines between group sections (except before the first group)
        if not first_group:
            blank = {col: '' for col in ordered_columns}
            manager_rows.extend([blank.copy(), blank.copy(), blank.copy()])
        first_group = False

        # (B) Group title row: e.g. "TEAM: TEAM A" in the 'Adviser' column
        # group_title_row = {col: '' for col in ordered_columns}
        # group_title_row['Adviser'] = f'TEAM: {str(team_group).upper()}'
        # manager_rows.append(group_title_row)

        # (C) Repeat column headers in the next row
        headers_row = {col: col for col in ordered_columns}
        manager_rows.append(headers_row)

        # (D) Compute subtotals for this group
        sum_issued_business  = grp_df['Issued Business $'].sum()
        sum_clawback_dollars = grp_df['Clawback $'].sum()
        sum_ntu_dollars      = grp_df['NTU $'].sum()
        sum_in_books         = grp_df['IN Books $'].sum()
        sum_issued_policies  = grp_df['Issued Policies Count'].sum()
        sum_clawback_count   = grp_df['Clawback Count'].sum()
        sum_ntu_count        = grp_df['NTU Count'].sum()
        sum_in_books_count   = grp_df['IN Books Count'].sum()

        # Team-level percentages
        if sum_issued_policies == 0:
            team_clawback_pct = 0
            team_ntu_pct      = 0
        else:
            team_clawback_pct = sum_clawback_count / sum_issued_policies
            team_ntu_pct      = sum_ntu_count / sum_issued_policies

        team_total_pct = team_clawback_pct + team_ntu_pct

        # (E) Subtotal row (goes first after the headers)
        # Here we multiply the percentages by 100 so 3% => 3.00, not 0.03
        subtotal_row = {
            'Adviser':               f'Team {str(team_group).upper()}:',
            'Issued Business $':     sum_issued_business,
            'Clawback $':            sum_clawback_dollars,
            'NTU $':                 sum_ntu_dollars,
            'IN Books $':            sum_in_books,
            'Issued Policies Count': sum_issued_policies,
            'Clawback Count':        sum_clawback_count,
            'NTU Count':             sum_ntu_count,
            'IN Books Count':        sum_in_books_count,
            'Clawback %':            round(team_clawback_pct*100, 2),
            'NTU %':                 round(team_ntu_pct*100, 2),
            'Total %':               round(team_total_pct*100, 2)
        }
        manager_rows.append(subtotal_row)

        # (F) Next, individual adviser rows in this Team Group
        for idx, detail_row in grp_df[ordered_columns].iterrows():
            manager_rows.append(detail_row.to_dict())

    # Finally, build a DataFrame from our accumulated row dicts:
    manager_report_df = pd.DataFrame(manager_rows, columns=ordered_columns)


    # =============================================================================
    # 8) EXCEL STYLING & COLOR CODING (FULL PARITY)
    # =============================================================================
    wb = Workbook()
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')
    thin_side = Side(border_style='thin')
    bold_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

    ws1 = wb.active
    ws1.title = "Clawback Report"
    for r in dataframe_to_rows(main_df, index=False, header=True): ws1.append(r)
    
    ws2 = wb.create_sheet("Business Mix")
    for r in dataframe_to_rows(df_pivot, index=False, header=True): ws2.append(r)
    
    ws3 = wb.create_sheet("Manager Report")
    for r in dataframe_to_rows(manager_report_df, index=False, header=False): ws3.append(r)

    # --- Insert Title into Sheet 3 FIRST to fix indexing ---
    raw_title = str(report_title_input).strip() if report_title_input else ""
    
    if not raw_title or raw_title.lower() in ["none", "null", "string"]:
        final_report_title = f"Clawback Report: {start_date.strftime('%b %Y')} to {end_date.strftime('%b %Y')}"
    else:
        final_report_title = raw_title
    # Find the first 'Adviser' cell to insert title above it
    for r_idx in range(1, ws3.max_row + 1):
        if ws3.cell(row=r_idx, column=1).value == "Adviser":
            ws3.insert_rows(r_idx)
            ws3.cell(row=r_idx, column=1, value=final_report_title).font = Font(bold=True, size=14)
            break

    # --- Sheet 3: Manager Report Section Highlighting (Fixed Syntax & Logic) ---
    
    
    row_idx = 1
    while row_idx <= ws3.max_row:
        cell_val = ws3.cell(row=row_idx, column=1).value
        
        if cell_val == final_report_title:
            row_idx += 1
            continue
        # If the cell is exactly "Adviser", it's a header row
        if isinstance(cell_val, str) and cell_val == "Adviser":
            for col in range(1, ws3.max_column + 1):
                # Header row styling
                ws3.cell(row=row_idx, column=col).font = bold_font
                ws3.cell(row=row_idx, column=col).border = bold_border
                ws3.cell(row=row_idx, column=col).fill = light_blue_fill
                # Subtotal row styling (the row immediately following header)
                ws3.cell(row=row_idx + 1, column=col).font = bold_font
                ws3.cell(row=row_idx + 1, column=col).border = bold_border
            row_idx += 2 # Move past header and subtotal
        else:
            row_idx += 1

    # Universal Formatting Helper (Borders, Centers, Colors)
    def apply_formatting(ws, apply_coloring=False):
        # We handle Sheet 3 headers differently because they repeat
        is_sheet_3 = (ws.title == "Manager Report")
        
        # Grab the header from the first row that actually has 'Adviser'
        header_row_vals = []
        for r in ws.iter_rows(min_row=1, max_row=5):
            vals = [c.value for c in r]
            if "Adviser" in vals:
                header_row_vals = vals
                break
        
        adviser_idx = next((i + 1 for i, v in enumerate(header_row_vals) if v == 'Adviser'), None)
        percent_cols = [i + 1 for i, v in enumerate(header_row_vals) if isinstance(v, str) and "%" in v]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Check if this specific row is a header row
            current_row_is_header = any(cell.value == "Adviser" for cell in row)
            
            for col_idx, cell in enumerate(row, 1):
                # Grid lines for Sheet 1 and 2
                if ws.title in ["Clawback Report", "Business Mix"]:
                    cell.border = bold_border
                
                # Adviser column bold
                if col_idx == adviser_idx and cell.value and not current_row_is_header:
                    cell.font = bold_font
                
                # Numbers centering
                if isinstance(cell.value, (int, float)):
                    cell.alignment = center_align
                
                # Conditional Coloring (10% yellow, 15% red)
                if apply_coloring and col_idx in percent_cols:
                    try:
                        val = float(cell.value)
                        if 10 <= val < 15: cell.fill = yellow_fill
                        elif val >= 15: cell.fill = red_fill
                    except: pass
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Apply to all sheets
    apply_formatting(ws1, apply_coloring=True)
    apply_formatting(ws2, apply_coloring=False)
    apply_formatting(ws3, apply_coloring=True)

    # Sheet 1: Manual TOTAL row logic (Final cleanup)
    headers = [cell.value for cell in ws1[1]]
    ws1.append([""] * len(headers)) # Blank spacer
    totals = {}
    for col_idx, h in enumerate(headers, 1):
        if h and "%" not in str(h) and h != "Adviser" and h != "Team Group":
            total_val = 0
            for r in list(ws1.iter_rows(min_row=2, max_row=ws1.max_row-1)):
                try: total_val += float(r[col_idx-1].value or 0)
                except: pass
            totals[h] = total_val
    
    issued_count = totals.get('Issued Policies Count', 0)
    totals['Clawback %'] = round(totals.get('Clawback Count', 0) / issued_count * 100, 2) if issued_count else 0
    totals['NTU %'] = round(totals.get('NTU Count', 0) / issued_count * 100, 2) if issued_count else 0
    totals['Total %'] = totals['Clawback %'] + totals['NTU %']
    
    final_total_row = ["TOTAL" if h == 'Adviser' else totals.get(h, 0) for h in headers]
    ws1.append(final_total_row)
    for cell in ws1[ws1.max_row]:
        cell.font, cell.fill, cell.border, cell.alignment = bold_font, blue_fill, bold_border, center_align

    # Final Save to Disk
    output_filename = f"clawback_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(tempfile.gettempdir(), output_filename)
    wb.save(output_path)
    return output_path, output_filename